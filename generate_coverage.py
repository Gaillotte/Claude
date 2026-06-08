"""
Generate PKCS11_Requirements_Coverage.xlsx
Sheets:
  1. Summary          – overall stats + pie-like bar summary
  2. Requirements     – one row per PKCS#11 v2.40 requirement with linked test IDs
  3. Test Results     – one row per test case with group / name / status / req IDs
  4. Coverage Matrix  – requirement × test group heatmap
"""

import re, subprocess, os
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.chart.series import DataPoint

# ──────────────────────────────────────────────────────────────────────────────
# Colour palette
# ──────────────────────────────────────────────────────────────────────────────
C_PASS   = "C6EFCE"   # green
C_FAIL   = "FFC7CE"   # red
C_SKIP   = "FFEB9C"   # yellow
C_HEAD1  = "1F4E79"   # dark blue
C_HEAD2  = "2E74B5"   # mid blue
C_HEAD3  = "BDD7EE"   # light blue
C_ALT    = "DEEAF1"   # very light blue
C_WHITE  = "FFFFFF"
C_TITLE  = "1F4E79"

def fill(hex_col):
    return PatternFill("solid", fgColor=hex_col)

def font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

thin = Side(style="thin", color="B8CCE4")
thick = Side(style="medium", color="2E74B5")

def thin_border():
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def thick_border():
    return Border(left=thick, right=thick, top=thick, bottom=thick)

# ──────────────────────────────────────────────────────────────────────────────
# Parse test results from the binary
# ──────────────────────────────────────────────────────────────────────────────
def run_tests():
    env = os.environ.copy()
    env["SOFTHSM2_CONF"] = os.path.expanduser("~/.softhsm2/softhsm2.conf")
    result = subprocess.run(
        ["/home/user/Claude/pkcs11_testsuite/build/pkcs11_test"],
        capture_output=True, text=True, env=env)
    raw = result.stdout + result.stderr
    # strip ANSI
    raw = re.sub(r'\x1b\[[0-9;]*m', '', raw)
    return raw

def parse_results(raw):
    tests = []
    for line in raw.splitlines():
        m_ok   = re.search(r'\[ OK \]\s+\[(\S+)\]\s+(.+)', line)
        m_fail = re.search(r'\[FAIL\]\s+\[(\S+)\]\s+(.+)', line)
        m_skip = re.search(r'SKIP\s+\[(\S+)\]\s+(.+)', line)
        if m_ok:
            tests.append({"group": m_ok.group(1),
                          "name":  m_ok.group(2).strip(),
                          "status":"PASS"})
        elif m_fail:
            tests.append({"group": m_fail.group(1),
                          "name":  m_fail.group(2).strip(),
                          "status":"FAIL"})
        elif m_skip:
            tests.append({"group": m_skip.group(1),
                          "name":  m_skip.group(2).strip(),
                          "status":"SKIP"})
    return tests

# ──────────────────────────────────────────────────────────────────────────────
# Requirements catalogue  (PKCS#11 v2.40 §11.x)
# ──────────────────────────────────────────────────────────────────────────────
REQUIREMENTS = [
    # id, section, function, description, [test_names...]
    # §11.4 General
    ("REQ-GEN-01","§11.4","C_Initialize",      "Accepts NULL args (uses defaults)",                                    ["C_Initialize_null","C_Initialize_os_locking"]),
    ("REQ-GEN-02","§11.4","C_Initialize",      "CKF_OS_LOCKING_OK enables OS primitives",                             ["C_Initialize_os_locking","Initialize_persistent_state"]),
    ("REQ-GEN-03","§11.4","C_Initialize",      "Double init → CKR_CRYPTOKI_ALREADY_INITIALIZED",                      ["C_Initialize_double"]),
    ("REQ-GEN-04","§11.4","C_Finalize",        "Normal finalization succeeds",                                         ["C_Initialize_null"]),
    ("REQ-GEN-05","§11.4","C_Finalize",        "Without init → CKR_CRYPTOKI_NOT_INITIALIZED",                         ["C_Finalize_without_init"]),
    ("REQ-GEN-06","§11.4","C_Finalize",        "pReserved != NULL → CKR_ARGUMENTS_BAD",                               ["REQ-GEN-01_Finalize_reserved_not_null"]),
    ("REQ-GEN-07","§11.4","C_GetInfo",         "Returns valid structure (version ≥ 2)",                               ["C_GetInfo","REQ-GEN-03_GetInfo_version"]),
    ("REQ-GEN-08","§11.4","C_GetInfo",         "Without init → CKR_CRYPTOKI_NOT_INITIALIZED",                         ["C_GetInfo_not_initialized"]),
    ("REQ-GEN-09","§11.4","C_GetInfo",         "NULL pInfo → CKR_ARGUMENTS_BAD",                                      ["C_GetInfo_null","REQ-GEN-02_GetInfo_null_ptr"]),
    ("REQ-GEN-10","§11.4","C_GetFunctionList", "Returns non-NULL list with all function pointers present",             ["C_GetFunctionList","REQ-GEN-05_GetFunctionList_consistent"]),
    ("REQ-GEN-11","§11.4","C_GetFunctionList", "NULL ppFunctionList → CKR_ARGUMENTS_BAD",                             ["C_GetFunctionList_null","REQ-GEN-04_GetFunctionList_null"]),
    # §11.5 Slot & Token
    ("REQ-SLOT-01","§11.5","C_GetSlotList",    "Two-call pattern: NULL then fill, counts consistent",                  ["C_GetSlotList_token_present","REQ-SLOT-02_GetSlotList_two_call"]),
    ("REQ-SLOT-02","§11.5","C_GetSlotList",    "tokenPresent=TRUE ≤ tokenPresent=FALSE (all slots)",                  ["C_GetSlotList_all","REQ-SLOT-03_GetSlotList_all_vs_token"]),
    ("REQ-SLOT-03","§11.5","C_GetSlotList",    "Buffer too small → CKR_BUFFER_TOO_SMALL",                             ["C_GetSlotList_buffer_too_small"]),
    ("REQ-SLOT-04","§11.5","C_GetSlotList",    "NULL count → CKR_ARGUMENTS_BAD",                                      ["REQ-SLOT-01_GetSlotList_null_count"]),
    ("REQ-SLOT-05","§11.5","C_GetSlotInfo",    "Returns valid CK_SLOT_INFO; CKF_TOKEN_PRESENT set",                   ["C_GetSlotInfo","REQ-SLOT-06_GetSlotInfo_token_present"]),
    ("REQ-SLOT-06","§11.5","C_GetSlotInfo",    "Invalid slot → CKR_SLOT_ID_INVALID",                                  ["C_GetSlotInfo_invalid","REQ-SLOT-05_GetSlotInfo_invalid_slot"]),
    ("REQ-SLOT-07","§11.5","C_GetSlotInfo",    "NULL pInfo → CKR_ARGUMENTS_BAD",                                      ["C_GetSlotInfo_null","REQ-SLOT-04_GetSlotInfo_null_ptr"]),
    ("REQ-SLOT-08","§11.5","C_GetTokenInfo",   "Returns valid CK_TOKEN_INFO; CKF_TOKEN_INITIALIZED set",              ["C_GetTokenInfo","REQ-SLOT-09_GetTokenInfo_initialized","REQ-SLOT-10_GetTokenInfo_field_bounds"]),
    ("REQ-SLOT-09","§11.5","C_GetTokenInfo",   "Invalid slot → CKR_SLOT_ID_INVALID",                                  ["C_GetTokenInfo_invalid_slot","REQ-SLOT-08_GetTokenInfo_invalid_slot"]),
    ("REQ-SLOT-10","§11.5","C_GetTokenInfo",   "NULL pInfo → CKR_ARGUMENTS_BAD",                                      ["C_GetTokenInfo_null","REQ-SLOT-07_GetTokenInfo_null_ptr"]),
    ("REQ-SLOT-11","§11.5","C_GetMechanismList","Two-call pattern; count > 0 for initialized token",                   ["C_GetMechanismList","C_GetMechanismList_buffer_too_small"]),
    ("REQ-SLOT-12","§11.5","C_GetMechanismList","Invalid slot → CKR_SLOT_ID_INVALID",                                  ["REQ-SLOT-12_GetMechList_invalid_slot"]),
    ("REQ-SLOT-13","§11.5","C_GetMechanismList","NULL count → CKR_ARGUMENTS_BAD",                                      ["REQ-SLOT-11_GetMechList_null_count"]),
    ("REQ-SLOT-14","§11.5","C_GetMechanismInfo","Returns flags indicating usage (ENCRYPT/DECRYPT/SIGN/VERIFY)",        ["C_GetMechanismInfo","REQ-SLOT-16_GetMechInfo_flags"]),
    ("REQ-SLOT-15","§11.5","C_GetMechanismInfo","Unknown mechanism → CKR_MECHANISM_INVALID",                           ["C_GetMechanismInfo_invalid","REQ-SLOT-15_GetMechInfo_unknown_mech"]),
    ("REQ-SLOT-16","§11.5","C_GetMechanismInfo","Invalid slot → CKR_SLOT_ID_INVALID",                                  ["REQ-SLOT-14_GetMechInfo_invalid_slot"]),
    ("REQ-SLOT-17","§11.5","C_GetMechanismInfo","NULL pInfo → CKR_ARGUMENTS_BAD",                                      ["REQ-SLOT-13_GetMechInfo_null_ptr"]),
    ("REQ-SLOT-18","§11.5","C_WaitForSlotEvent","Non-blocking: returns CKR_OK or CKR_NO_EVENT immediately",            ["C_WaitForSlotEvent_no_block","REQ-SLOT-17_WaitForSlotEvent_nonblocking"]),
    ("REQ-SLOT-19","§11.5","C_WaitForSlotEvent","NULL pSlot: CKR_ARGUMENTS_BAD or CKR_NO_EVENT",                      ["REQ-SLOT-18_WaitForSlotEvent_null"]),
    ("REQ-SLOT-20","§11.5","C_InitPIN",        "Requires SO session (RW_SO_FUNCTIONS); USER session → rejected",       ["REQ-SLOT-20_InitPIN_wrong_state"]),
    ("REQ-SLOT-21","§11.5","C_SetPIN",         "Wrong old PIN → CKR_PIN_INCORRECT",                                    ["REQ-SLOT-19_SetPIN_wrong_old_pin","SetPIN"]),
    # §11.6 Session management
    ("REQ-SESS-01","§11.6","C_OpenSession",    "CKF_SERIAL_SESSION must be set; omit → CKR_SESSION_PARALLEL_NOT_SUPPORTED",["REQ-SESS-01_OpenSession_no_serial_flag"]),
    ("REQ-SESS-02","§11.6","C_OpenSession",    "Invalid slot → CKR_SLOT_ID_INVALID",                                   ["REQ-SESS-02_OpenSession_invalid_slot"]),
    ("REQ-SESS-03","§11.6","C_OpenSession",    "NULL phSession → CKR_ARGUMENTS_BAD",                                   ["REQ-SESS-03_OpenSession_null_handle"]),
    ("REQ-SESS-04","§11.6","C_OpenSession",    "CKF_RW_SESSION produces RW session (CKS_RW_*)",                        ["OpenSession_RW","REQ-SESS-08_SessionState_rw_user","REQ-SESS-10_SessionInfo_slot"]),
    ("REQ-SESS-05","§11.6","C_OpenSession",    "Without flag opens RO session (CKS_RO_*)",                             ["OpenSession_RO","REQ-SESS-09_SessionState_ro_user"]),
    ("REQ-SESS-06","§11.6","C_CloseSession",   "Invalid handle → CKR_SESSION_HANDLE_INVALID",                          ["CloseSession_invalid_handle","REQ-SESS-04_CloseSession_invalid_handle"]),
    ("REQ-SESS-07","§11.6","C_CloseSession",   "Active operations aborted on close",                                   ["REQ-SESS-11_SessionObjects_destroyed_on_close"]),
    ("REQ-SESS-08","§11.6","C_CloseSession",   "Session objects destroyed when session closes",                        ["REQ-SESS-11_SessionObjects_destroyed_on_close"]),
    ("REQ-SESS-09","§11.6","C_CloseAllSessions","Closes all sessions on slot",                                         ["CloseAllSessions"]),
    ("REQ-SESS-10","§11.6","C_CloseAllSessions","Invalid slot → CKR_SLOT_ID_INVALID",                                  ["REQ-SESS-05_CloseAllSessions_invalid_slot"]),
    ("REQ-SESS-11","§11.6","C_GetSessionInfo", "Invalid handle → CKR_SESSION_HANDLE_INVALID",                          ["GetSessionInfo_invalid_handle","REQ-SESS-06_GetSessionInfo_invalid_handle"]),
    ("REQ-SESS-12","§11.6","C_GetSessionInfo", "NULL pInfo → CKR_ARGUMENTS_BAD",                                       ["REQ-SESS-07_GetSessionInfo_null"]),
    ("REQ-SESS-13","§11.6","C_GetSessionInfo", "state and flags reflect login status and session type",                 ["REQ-SESS-08_SessionState_rw_user","REQ-SESS-09_SessionState_ro_user","REQ-SESS-10_SessionInfo_slot"]),
    ("REQ-SESS-14","§11.6","C_GetOperationState","NULL buffer returns required length or NOT_INITIALIZED",              ["GetOperationState_idle","REQ-SESS-12_GetOperationState_size_query"]),
    ("REQ-SESS-15","§11.6","C_GetOperationState","NULL length → CKR_ARGUMENTS_BAD",                                    ["REQ-SESS-13_GetOperationState_null_len"]),
    # §11.6 Login / Logout
    ("REQ-LOG-01","§11.6","C_Login",           "CKU_USER with correct PIN succeeds",                                   ["Login_user"]),
    ("REQ-LOG-02","§11.6","C_Login",           "CKU_SO with correct SO PIN succeeds",                                  ["Login_so"]),
    ("REQ-LOG-03","§11.6","C_Login",           "Wrong PIN → CKR_PIN_INCORRECT",                                        ["Login_wrong_pin","REQ-LOG-02_Login_wrong_pin"]),
    ("REQ-LOG-04","§11.6","C_Login",           "Already logged in → CKR_USER_ALREADY_LOGGED_IN",                       ["Login_already_logged_in","REQ-LOG-01_Login_already_logged_in"]),
    ("REQ-LOG-05","§11.6","C_Login",           "SO while USER logged in → blocked (ANOTHER_LOGGED_IN or READ_ONLY_EXISTS)",["REQ-LOG-04_Login_SO_while_user_logged_in"]),
    ("REQ-LOG-06","§11.6","C_Login",           "Invalid user type → CKR_USER_TYPE_INVALID",                            ["REQ-LOG-03_Login_invalid_user_type"]),
    ("REQ-LOG-07","§11.6","C_Login",           "Login state shared across all sessions on slot",                       ["REQ-LOG-07_Login_state_shared"]),
    ("REQ-LOG-08","§11.6","C_Login",           "Zero-length PIN rejected",                                             ["REQ-LOG-08_Login_zero_length_pin"]),
    ("REQ-LOG-09","§11.6","C_Login",           "NULL PIN with non-zero length → error",                               ["REQ-LOG-09_Login_null_pin"]),
    ("REQ-LOG-10","§11.6","C_Login",           "SO login on RO session rejected",                                      ["REQ-LOG-10_SO_login_rejected_on_ro_session"]),
    ("REQ-LOG-11","§11.6","C_Logout",          "Returns all sessions to public state",                                  ["REQ-LOG-05_Logout_reverts_state"]),
    ("REQ-LOG-12","§11.6","C_Logout",          "Not logged in → CKR_USER_NOT_LOGGED_IN or OK",                        ["Logout_not_logged_in","REQ-LOG-06_Logout_when_not_logged_in"]),
    # §11.7 Objects
    ("REQ-OBJ-01","§11.7","C_CreateObject",    "Missing required attribute → CKR_TEMPLATE_INCOMPLETE",                 ["REQ-OBJ-01_CreateObject_missing_class"]),
    ("REQ-OBJ-02","§11.7","C_CreateObject",    "NULL template with count > 0 → CKR_ARGUMENTS_BAD",                   ["REQ-OBJ-02_CreateObject_null_template"]),
    ("REQ-OBJ-03","§11.7","C_CreateObject",    "NULL phObject → CKR_ARGUMENTS_BAD",                                   ["REQ-OBJ-03_CreateObject_null_handle"]),
    ("REQ-OBJ-04","§11.7","C_CreateObject",    "token=TRUE on RO session → CKR_SESSION_READ_ONLY",                    ["REQ-OBJ-04_CreateObject_token_on_ro_session"]),
    ("REQ-OBJ-05","§11.7","C_CreateObject",    "Token object visible across sessions",                                 ["REQ-OBJ-05_Token_object_persists"]),
    ("REQ-OBJ-06","§11.7","C_CreateObject",    "Session object destroyed when session closes",                         ["REQ-SESS-11_SessionObjects_destroyed_on_close"]),
    ("REQ-OBJ-07","§11.7","C_CreateObject",    "CKO_DATA object creation",                                             ["CreateObject_data","REQ-OBJ-26_CreateObject_certificate","REQ-OBJ-27_CreateObject_secret_key_value"]),
    ("REQ-OBJ-08","§11.7","C_CopyObject",      "Creates independent copy with modified attributes",                    ["CopyObject","REQ-OBJ-06_CopyObject_independent"]),
    ("REQ-OBJ-09","§11.7","C_CopyObject",      "NULL phNewObject → CKR_ARGUMENTS_BAD",                               ["REQ-OBJ-07_CopyObject_null_handle"]),
    ("REQ-OBJ-10","§11.7","C_DestroyObject",   "Invalid handle → CKR_OBJECT_HANDLE_INVALID",                          ["DestroyObject_invalid","REQ-OBJ-08_DestroyObject_invalid_handle"]),
    ("REQ-OBJ-11","§11.7","C_GetObjectSize",   "Returns non-zero size or CKR_INFORMATION_SENSITIVE",                  ["GetObjectSize","REQ-OBJ-11_GetObjectSize_valid"]),
    ("REQ-OBJ-12","§11.7","C_GetObjectSize",   "Invalid handle → CKR_OBJECT_HANDLE_INVALID",                          ["REQ-OBJ-09_GetObjectSize_invalid"]),
    ("REQ-OBJ-13","§11.7","C_GetObjectSize",   "NULL pulSize → CKR_ARGUMENTS_BAD",                                    ["REQ-OBJ-10_GetObjectSize_null"]),
    ("REQ-OBJ-14","§11.7","C_GetAttributeValue","NULL buffer → size query; ulValueLen set to required",               ["GetAttributeValue","REQ-OBJ-12_GetAttr_null_buffer_size_query"]),
    ("REQ-OBJ-15","§11.7","C_GetAttributeValue","Buffer too small → CKR_BUFFER_TOO_SMALL",                            ["GetAttr_buffer_too_small","REQ-OBJ-15_GetAttr_buffer_too_small"]),
    ("REQ-OBJ-16","§11.7","C_GetAttributeValue","Sensitive attribute → CKR_ATTRIBUTE_SENSITIVE / unavailable",        ["REQ-OBJ-14_GetAttr_sensitive","SensitiveAttribute"]),
    ("REQ-OBJ-17","§11.7","C_GetAttributeValue","Non-existent attribute → CKR_ATTRIBUTE_TYPE_INVALID",                ["REQ-OBJ-13_GetAttr_nonexistent"]),
    ("REQ-OBJ-18","§11.7","C_GetAttributeValue","Invalid handle → CKR_OBJECT_HANDLE_INVALID",                         ["GetAttrInvalidObject","REQ-OBJ-16_GetAttr_invalid_object"]),
    ("REQ-OBJ-19","§11.7","C_SetAttributeValue","Read-only attribute → CKR_ATTRIBUTE_READ_ONLY",                      ["SetAttributeValue","REQ-OBJ-17_SetAttr_read_only"]),
    ("REQ-OBJ-20","§11.7","C_SetAttributeValue","RO session → CKR_SESSION_READ_ONLY",                                 ["REQ-OBJ-18_SetAttr_ro_session"]),
    ("REQ-OBJ-21","§11.7","C_SetAttributeValue","Invalid handle → CKR_OBJECT_HANDLE_INVALID",                         ["REQ-OBJ-19_SetAttr_invalid_object"]),
    ("REQ-OBJ-22","§11.7","C_FindObjectsInit", "Second Init while active → CKR_OPERATION_ACTIVE",                     ["FindObjects_active_op","REQ-OBJ-20_FindObjectsInit_while_active"]),
    ("REQ-OBJ-23","§11.7","C_FindObjects",     "Without Init → CKR_OPERATION_NOT_INITIALIZED",                        ["REQ-OBJ-21_FindObjects_without_init"]),
    ("REQ-OBJ-24","§11.7","C_FindObjects",     "Returns at most ulMaxObjectCount handles",                             ["REQ-OBJ-24_FindObjects_max_count"]),
    ("REQ-OBJ-25","§11.7","C_FindObjects",     "NULL phObject → CKR_ARGUMENTS_BAD",                                   ["REQ-OBJ-23_FindObjects_null_handle"]),
    ("REQ-OBJ-26","§11.7","C_FindObjects",     "Multi-attribute template filters correctly",                           ["FindObjects","REQ-OBJ-25_FindObjects_multi_attr_filter"]),
    ("REQ-OBJ-27","§11.7","C_FindObjectsFinal","Without Init → CKR_OPERATION_NOT_INITIALIZED",                        ["REQ-OBJ-22_FindObjectsFinal_without_init"]),
    # §11.8 Encryption
    ("REQ-ENC-01","§11.8","C_EncryptInit",     "Invalid mechanism → CKR_MECHANISM_INVALID",                           ["REQ-ENC-04_EncryptInit_invalid_mechanism"]),
    ("REQ-ENC-02","§11.8","C_EncryptInit",     "NULL mechanism → CKR_ARGUMENTS_BAD",                                  ["REQ-ENC-06_EncryptInit_null_mechanism"]),
    ("REQ-ENC-03","§11.8","C_EncryptInit",     "Key without CKA_ENCRYPT → CKR_KEY_FUNCTION_NOT_PERMITTED",            ["REQ-ENC-05_EncryptInit_key_no_encrypt"]),
    ("REQ-ENC-04","§11.8","C_EncryptInit",     "Key type mismatch → CKR_KEY_TYPE_INCONSISTENT",                       ["REQ-ENC-09_Encrypt_wrong_key_type","DecryptInit_wrong_key_type"]),
    ("REQ-ENC-05","§11.8","C_EncryptInit",     "While op active → CKR_OPERATION_ACTIVE",                              ["EncryptInit_reuse","REQ-ENC-07_EncryptInit_while_active"]),
    ("REQ-ENC-06","§11.8","C_Encrypt",         "Without Init → CKR_OPERATION_NOT_INITIALIZED",                        ["EncryptNoInit","REQ-ENC-01_Encrypt_without_init"]),
    ("REQ-ENC-07","§11.8","C_Encrypt",         "NULL output buffer → size query returns required length",             ["Encrypt_get_output_length","REQ-ENC-08_Encrypt_null_output_query"]),
    ("REQ-ENC-08","§11.8","C_Encrypt",         "Buffer too small → CKR_BUFFER_TOO_SMALL",                             ["Encrypt_buffer_too_small"]),
    ("REQ-ENC-09","§11.8","C_Encrypt",         "AES-CBC-PAD encrypt+decrypt roundtrip",                               ["AES_128_CBC_PAD","AES_192_CBC_PAD","AES_256_CBC_PAD"]),
    ("REQ-ENC-10","§11.8","C_Encrypt",         "AES-ECB encrypt+decrypt roundtrip",                                   ["AES_ECB_roundtrip"]),
    ("REQ-ENC-11","§11.8","C_Encrypt",         "AES-CTR encrypt+decrypt roundtrip",                                   ["AES_CTR_roundtrip"]),
    ("REQ-ENC-12","§11.8","C_Encrypt",         "AES-GCM encrypt+decrypt roundtrip with AAD",                          ["AES_GCM_roundtrip"]),
    ("REQ-ENC-13","§11.8","C_EncryptUpdate",   "Without Init → CKR_OPERATION_NOT_INITIALIZED",                        ["EncryptUpdateNoInit","REQ-ENC-02_EncryptUpdate_without_init"]),
    ("REQ-ENC-14","§11.8","C_EncryptUpdate",   "Multi-part encrypt produces same result",                              ["AES_multipart_encrypt"]),
    ("REQ-ENC-15","§11.8","C_EncryptFinal",    "Without Init → CKR_OPERATION_NOT_INITIALIZED",                        ["REQ-ENC-03_EncryptFinal_without_init"]),
    # §11.9 Decryption
    ("REQ-DEC-01","§11.9","C_DecryptInit",     "Key without CKA_DECRYPT → CKR_KEY_FUNCTION_NOT_PERMITTED",            ["REQ-DEC-04_DecryptInit_key_no_decrypt"]),
    ("REQ-DEC-02","§11.9","C_Decrypt",         "Without Init → CKR_OPERATION_NOT_INITIALIZED",                        ["REQ-DEC-01_Decrypt_without_init"]),
    ("REQ-DEC-03","§11.9","C_Decrypt",         "Corrupt ciphertext → CKR_ENCRYPTED_DATA_INVALID",                     ["DecryptWrongKey","REQ-DEC-05_Decrypt_corrupt_ciphertext"]),
    ("REQ-DEC-04","§11.9","C_DecryptUpdate",   "Without Init → CKR_OPERATION_NOT_INITIALIZED",                        ["REQ-DEC-02_DecryptUpdate_without_init"]),
    ("REQ-DEC-05","§11.9","C_DecryptUpdate",   "Multi-part decrypt produces same result",                              ["AES_multipart_decrypt"]),
    ("REQ-DEC-06","§11.9","C_DecryptFinal",    "Without Init → CKR_OPERATION_NOT_INITIALIZED",                        ["REQ-DEC-03_DecryptFinal_without_init"]),
    # §11.10 Digest
    ("REQ-DIG-01","§11.10","C_DigestInit",     "While active → CKR_OPERATION_ACTIVE",                                 ["REQ-DIG-04_DigestInit_while_active"]),
    ("REQ-DIG-02","§11.10","C_Digest",         "Without Init → CKR_OPERATION_NOT_INITIALIZED",                        ["REQ-DIG-01_Digest_without_init"]),
    ("REQ-DIG-03","§11.10","C_Digest",         "NULL output → size query",                                            ["REQ-DIG-06_Digest_null_output_query"]),
    ("REQ-DIG-04","§11.10","C_Digest",         "SHA-1/256/384/512 produce correct digest lengths",                    ["SHA1_digest","SHA256_digest","SHA384_digest","SHA512_digest"]),
    ("REQ-DIG-05","§11.10","C_DigestUpdate",   "Without Init → CKR_OPERATION_NOT_INITIALIZED",                        ["REQ-DIG-02_DigestUpdate_without_init"]),
    ("REQ-DIG-06","§11.10","C_DigestUpdate",   "Multi-part equals single-call",                                       ["Digest_multipart","REQ-DIG-07_Digest_multipart_matches_single"]),
    ("REQ-DIG-07","§11.10","C_DigestFinal",    "Without Init → CKR_OPERATION_NOT_INITIALIZED",                        ["REQ-DIG-03_DigestFinal_without_init"]),
    ("REQ-DIG-08","§11.10","C_DigestFinal",    "Buffer too small → CKR_BUFFER_TOO_SMALL",                             ["Digest_buffer_too_small"]),
    ("REQ-DIG-09","§11.10","C_DigestKey",      "Incorporates secret key material into digest",                        ["REQ-DIG-05_DigestKey"]),
    # §11.11 Sign
    ("REQ-SIG-01","§11.11","C_SignInit",       "Key without CKA_SIGN → CKR_KEY_FUNCTION_NOT_PERMITTED",               ["REQ-SIG-04_SignInit_key_no_sign"]),
    ("REQ-SIG-02","§11.11","C_Sign",           "Without Init → CKR_OPERATION_NOT_INITIALIZED",                        ["REQ-SIG-01_Sign_without_init"]),
    ("REQ-SIG-03","§11.11","C_Sign",           "NULL sig buffer → size query",                                        ["REQ-SIG-05_Sign_null_output_query"]),
    ("REQ-SIG-04","§11.11","C_Sign",           "RSA PKCS#1 v1.5 sign/verify",                                         ["RSA_PKCS1_sign_verify"]),
    ("REQ-SIG-05","§11.11","C_Sign",           "RSA-PSS sign/verify",                                                 ["RSA_PSS_sign_verify"]),
    ("REQ-SIG-06","§11.11","C_SignUpdate",     "Without Init → CKR_OPERATION_NOT_INITIALIZED",                        ["REQ-SIG-02_SignUpdate_without_init"]),
    ("REQ-SIG-07","§11.11","C_SignUpdate",     "Multi-part sign matches single-call",                                  ["REQ-SIG-06_Sign_multipart"]),
    ("REQ-SIG-08","§11.11","C_SignFinal",      "Without Init → CKR_OPERATION_NOT_INITIALIZED",                        ["REQ-SIG-03_SignFinal_without_init"]),
    # §11.12 Verify
    ("REQ-VER-01","§11.12","C_VerifyInit",     "Key without CKA_VERIFY → CKR_KEY_FUNCTION_NOT_PERMITTED",             ["REQ-VER-06_VerifyInit_key_no_verify"]),
    ("REQ-VER-02","§11.12","C_Verify",         "Without Init → CKR_OPERATION_NOT_INITIALIZED",                        ["REQ-VER-01_Verify_without_init"]),
    ("REQ-VER-03","§11.12","C_Verify",         "Invalid signature → CKR_SIGNATURE_INVALID",                           ["TamperedSign","RSA_tampered_signature","REQ-VER-04_Verify_invalid_signature"]),
    ("REQ-VER-04","§11.12","C_Verify",         "Wrong length → CKR_SIGNATURE_LEN_RANGE or SIGNATURE_INVALID",         ["REQ-VER-05_Verify_wrong_sig_length"]),
    ("REQ-VER-05","§11.12","C_VerifyUpdate",   "Without Init → CKR_OPERATION_NOT_INITIALIZED",                        ["REQ-VER-02_VerifyUpdate_without_init"]),
    ("REQ-VER-06","§11.12","C_VerifyUpdate",   "Multi-part verify matches single-call",                               ["REQ-VER-07_Verify_multipart"]),
    ("REQ-VER-07","§11.12","C_VerifyFinal",    "Without Init → CKR_OPERATION_NOT_INITIALIZED",                        ["REQ-VER-03_VerifyFinal_without_init"]),
    # §11.14 Key management
    ("REQ-KEY-01","§11.14","C_GenerateKey",    "NULL mechanism → CKR_ARGUMENTS_BAD",                                  ["REQ-KEY-01_GenerateKey_null_mechanism"]),
    ("REQ-KEY-02","§11.14","C_GenerateKey",    "NULL phKey → CKR_ARGUMENTS_BAD",                                      ["REQ-KEY-02_GenerateKey_null_handle"]),
    ("REQ-KEY-03","§11.14","C_GenerateKey",    "Invalid mechanism → CKR_MECHANISM_INVALID",                           ["REQ-KEY-03_GenerateKey_invalid_mechanism"]),
    ("REQ-KEY-04","§11.14","C_GenerateKey",    "AES-128/192/256 key generation",                                      ["AES_128_keygen","AES_192_keygen","AES_256_keygen"]),
    ("REQ-KEY-05","§11.14","C_GenerateKeyPair","NULL mechanism → CKR_ARGUMENTS_BAD",                                  ["REQ-KEY-04_GenerateKeyPair_null_mechanism"]),
    ("REQ-KEY-06","§11.14","C_GenerateKeyPair","RSA-1024/2048/4096 key pair generation",                              ["RSA_1024_keygen","RSA_2048_keygen"]),
    ("REQ-KEY-07","§11.14","C_GenerateKeyPair","EC P-256/P-384/P-521 key pair generation",                            ["EC_P256_keygen","EC_P384_keygen","EC_P521_keygen"]),
    ("REQ-KEY-08","§11.14","C_WrapKey",        "Non-extractable key → CKR_KEY_UNEXTRACTABLE",                         ["WrapNonExtractable","REQ-KEY-05_WrapKey_non_extractable"]),
    ("REQ-KEY-09","§11.14","C_WrapKey",        "NULL wrapped buffer → size query",                                    ["REQ-KEY-06_WrapKey_null_output_query"]),
    ("REQ-KEY-10","§11.14","C_WrapKey",        "Wrap+unwrap roundtrip produces usable key",                           ["AES_wrap_unwrap","REQ-KEY-07_UnwrapKey_attributes"]),
    ("REQ-KEY-11","§11.14","C_UnwrapKey",      "Unwrapped key usable for encrypt/decrypt",                            ["AES_unwrap_verify","REQ-KEY-07_UnwrapKey_attributes"]),
    ("REQ-KEY-12","§11.14","C_DeriveKey",      "NULL mechanism → CKR_ARGUMENTS_BAD",                                  ["REQ-KEY-08_DeriveKey_null_mechanism"]),
    ("REQ-KEY-13","§11.14","C_DeriveKey",      "ECDH key derivation produces usable key",                             ["ECDH1_derive","ECDH_both_parties"]),
    # §11.15 Random
    ("REQ-RND-01","§11.15","C_GenerateRandom","Zero length → CKR_OK",                                                 ["GenerateRandom_zero","REQ-RND-01_Random_zero_length"]),
    ("REQ-RND-02","§11.15","C_GenerateRandom","NULL buffer non-zero length → CKR_ARGUMENTS_BAD",                      ["REQ-RND-02_Random_null_buffer"]),
    ("REQ-RND-03","§11.15","C_GenerateRandom","Output is non-constant (random)",                                      ["GenerateRandom_32","GenerateRandom_large","REQ-RND-03_Random_non_constant"]),
    ("REQ-RND-04","§11.15","C_SeedRandom",    "NULL buffer → CKR_ARGUMENTS_BAD or SEED_NOT_SUPPORTED",               ["SeedRandom","REQ-RND-04_SeedRandom_null_buffer"]),
    ("REQ-RND-05","§11.15","C_SeedRandom",    "Valid seed accepted or CKR_RANDOM_SEED_NOT_SUPPORTED",                 ["REQ-RND-05_SeedRandom_valid"]),
    # §11.13 Parallel / Dual
    ("REQ-PAR-01","§11.x","Parallel",         "4 concurrent sessions: AES encrypt/decrypt, no corruption",            ["Parallel_encrypt_decrypt"]),
    ("REQ-PAR-02","§11.x","Parallel",         "4 concurrent sessions: SHA-256 digest, no corruption",                ["Parallel_digest"]),
    ("REQ-DUL-01","§11.13","C_DigestEncryptUpdate","Dual-function op (skip if not supported)",                        ["DigestEncryptUpdate"]),
    ("REQ-DUL-02","§11.13","C_DecryptDigestUpdate","Dual-function op (skip if not supported)",                        ["DecryptDigestUpdate"]),
    ("REQ-DUL-03","§11.13","C_SignEncryptUpdate",  "Dual-function op (skip if not supported)",                        ["SignEncryptUpdate"]),
    # Negative / error-path scenarios
    ("REQ-NEG-01","§11.x","All",                 "Invalid session handle → CKR_SESSION_HANDLE_INVALID",               ["NEG-SESS-01_EncryptInit_invalid_sess","NEG-SESS-02_DecryptInit_invalid_sess","NEG-SESS-03_DigestInit_invalid_sess","NEG-SESS-04_SignInit_invalid_sess","NEG-SESS-05_VerifyInit_invalid_sess","NEG-SESS-06_GenerateKey_invalid_sess","NEG-SESS-07_GenerateKeyPair_invalid_sess","NEG-SESS-08_WrapKey_invalid_sess","NEG-SESS-09_FindObjectsInit_invalid_sess","NEG-SESS-10_CreateObject_invalid_sess","NEG-SESS-11_DestroyObject_invalid_sess","NEG-SESS-12_GenerateRandom_invalid_sess","NEG-SESS-13_Login_invalid_sess","NEG-SESS-14_Logout_invalid_sess","NEG-SESS-15_GetAttributeValue_invalid_sess"]),
    ("REQ-NEG-02","§11.x","C_EncryptInit/C_DecryptInit/C_SignInit/C_VerifyInit","Invalid key handle → CKR_KEY_HANDLE_INVALID or CKR_OBJECT_HANDLE_INVALID",["NEG-KEY-01_EncryptInit_invalid_key","NEG-KEY-02_DecryptInit_invalid_key","NEG-KEY-03_SignInit_invalid_key","NEG-KEY-04_VerifyInit_invalid_key"]),
    ("REQ-NEG-03","§11.x","C_WrapKey",           "Invalid wrapping key → CKR_WRAPPING_KEY_HANDLE_INVALID",            ["NEG-KEY-05_WrapKey_invalid_wrapping","NEG-KEY-06_WrapKey_invalid_target"]),
    ("REQ-NEG-04","§11.x","C_SignInit",          "Wrong key type → CKR_KEY_TYPE_INCONSISTENT (or deferred error)",    ["NEG-KEY-07_Sign_wrong_key_type","NEG-MECH-08_RSA_sign_mech_with_AES_key"]),
    ("REQ-NEG-05","§11.x","C_EncryptInit",       "Wrong key type/function → CKR_KEY_FUNCTION_NOT_PERMITTED",         ["NEG-KEY-08_Encrypt_rsa_key_for_aes"]),
    ("REQ-NEG-06","§11.x","C_EncryptInit",       "NULL IV for CBC → CKR_MECHANISM_PARAM_INVALID or CKR_ARGUMENTS_BAD",["NEG-MECH-01_CBC_null_iv","NEG-MECH-02_CBC_wrong_iv_len","NEG-MECH-03_CBC_zero_iv_len"]),
    ("REQ-NEG-07","§11.x","C_EncryptInit",       "NULL GCM params → error",                                          ["NEG-MECH-04_GCM_null_params"]),
    ("REQ-NEG-08","§11.x","C_Encrypt",           "Data too long for mechanism → CKR_DATA_LEN_RANGE",                 ["NEG-MECH-05_RSA_data_too_long"]),
    ("REQ-NEG-09","§11.x","C_Decrypt",           "Non-block-aligned ciphertext → CKR_ENCRYPTED_DATA_LEN_RANGE",      ["NEG-MECH-06_CBC_unaligned_decrypt"]),
    ("REQ-NEG-10","§11.x","C_Decrypt",           "Invalid padding → CKR_ENCRYPTED_DATA_INVALID",                     ["NEG-MECH-07_CBCPAD_bad_padding"]),
    ("REQ-NEG-11","§11.x","C_Encrypt/C_Decrypt", "NULL input data non-zero length → CKR_ARGUMENTS_BAD",             ["NEG-DATA-01_Encrypt_null_input","NEG-DATA-02_Decrypt_null_input"]),
    ("REQ-NEG-12","§11.x","C_DigestUpdate",      "NULL input → CKR_ARGUMENTS_BAD",                                  ["NEG-DATA-03_Digest_null_input"]),
    ("REQ-NEG-13","§11.x","C_Sign/C_Verify",     "NULL input/sig → CKR_ARGUMENTS_BAD",                              ["NEG-DATA-04_Sign_null_input","NEG-DATA-05_Verify_null_input","NEG-DATA-06_Verify_null_sig"]),
    ("REQ-NEG-14","§11.x","C_Encrypt",           "NULL output length → CKR_ARGUMENTS_BAD",                          ["NEG-DATA-07_Encrypt_null_outlen"]),
    ("REQ-NEG-15","§11.x","C_EncryptFinal/C_DecryptFinal/C_DigestFinal/C_SignFinal","Buffer too small → CKR_BUFFER_TOO_SMALL",["NEG-SEQ-01_EncryptFinal_too_small","NEG-SEQ-02_DecryptFinal_too_small","NEG-SEQ-03_DigestFinal_too_small","NEG-SEQ-04_SignFinal_too_small"]),
    ("REQ-NEG-16","§11.x","C_Encrypt/C_EncryptUpdate","Mixing single/multi-part ops",                                ["NEG-SEQ-05_Encrypt_after_EncryptUpdate","NEG-SEQ-06_EncryptUpdate_after_Encrypt"]),
    ("REQ-NEG-17","§11.x","C_SignFinal/C_VerifyFinal","Final without Update (empty message or error)",               ["NEG-SEQ-07_SignFinal_no_update","NEG-SEQ-08_VerifyFinal_no_update"]),
    ("REQ-NEG-18","§11.x","C_DestroyObject",     "RO session cannot destroy token objects → CKR_SESSION_READ_ONLY",  ["NEG-OBJ-01_Destroy_on_ro_session"]),
    ("REQ-NEG-19","§11.x","C_CopyObject",        "Invalid source handle → CKR_OBJECT_HANDLE_INVALID",               ["NEG-OBJ-02_CopyObject_invalid_src"]),
    ("REQ-NEG-20","§11.x","C_DestroyObject",     "CKA_DESTROYABLE=FALSE → CKR_ACTION_PROHIBITED",                   ["NEG-OBJ-03_Non_destroyable_object"]),
    ("REQ-NEG-21","§11.x","C_SetAttributeValue/C_GetAttributeValue","NULL template non-zero count → CKR_ARGUMENTS_BAD",["NEG-OBJ-04_SetAttr_null_template","NEG-OBJ-05_GetAttr_null_template"]),
    ("REQ-NEG-22","§11.x","C_FindObjectsInit",   "NULL template non-zero count → CKR_ARGUMENTS_BAD",               ["NEG-OBJ-06_FindInit_null_tmpl_nonzero"]),
    ("REQ-NEG-23","§11.x","C_CreateObject",      "Inconsistent template → CKR_TEMPLATE_INCONSISTENT",              ["NEG-OBJ-07_CreateObject_inconsistent_tmpl"]),
    ("REQ-NEG-24","§11.x","C_WrapKey",           "Invalid/unsupported wrap mechanism → error",                      ["NEG-WRAP-01_WrapKey_invalid_mechanism"]),
    ("REQ-NEG-25","§11.x","C_UnwrapKey",         "Corrupt wrapped data → CKR_WRAPPED_KEY_INVALID",                  ["NEG-WRAP-02_UnwrapKey_corrupt_blob"]),
    ("REQ-NEG-26","§11.x","C_UnwrapKey",         "NULL wrapped data → CKR_ARGUMENTS_BAD",                          ["NEG-WRAP-03_UnwrapKey_null_data"]),
    ("REQ-NEG-27","§11.x","C_WrapKey",           "Key without CKA_WRAP → CKR_KEY_NOT_WRAPPABLE",                   ["NEG-WRAP-04_WrapKey_no_wrap_flag"]),
    ("REQ-NEG-28","§11.x","C_GenerateKey",       "Invalid key length → CKR_KEY_SIZE_RANGE",                        ["NEG-KGEN-01_AES_invalid_length"]),
    ("REQ-NEG-29","§11.x","C_GenerateKey",       "Missing VALUE_LEN → CKR_TEMPLATE_INCOMPLETE",                    ["NEG-KGEN-02_AES_missing_value_len"]),
    ("REQ-NEG-30","§11.x","C_GenerateKeyPair",   "Invalid RSA modulus bits → error",                               ["NEG-KGEN-03_RSA_zero_bits"]),
    ("REQ-NEG-31","§11.x","C_GenerateKey",       "CKA_TOKEN=TRUE on RO session → CKR_SESSION_READ_ONLY",           ["NEG-KGEN-04_GenerateKey_on_ro_session"]),
    ("REQ-NEG-32","§11.x","C_DecryptInit/C_SignInit","Second init while op active → CKR_OPERATION_ACTIVE",          ["NEG-ENC-01_DecryptInit_while_Encrypt_active","NEG-ENC-02_SignInit_while_Digest_active"]),
    ("REQ-NEG-33","§11.x","C_Decrypt",           "GCM authentication tag tampered → CKR_ENCRYPTED_DATA_INVALID",   ["NEG-ENC-03_GCM_tampered_tag"]),
    ("REQ-NEG-34","§11.x","C_Decrypt",           "Random RSA ciphertext: implementation-defined behavior",         ["NEG-ENC-04_RSA_random_ciphertext"]),
    ("REQ-NEG-35","§11.x","C_Decrypt",           "Buffer too small → CKR_BUFFER_TOO_SMALL",                        ["NEG-ENC-05_Decrypt_buffer_too_small"]),
]

# ──────────────────────────────────────────────────────────────────────────────
# Build test index from results
# ──────────────────────────────────────────────────────────────────────────────
print("Running tests …")
raw = run_tests()
tests = parse_results(raw)
test_index = {t["name"]: t for t in tests}

print(f"  Parsed {len(tests)} test results")

# ──────────────────────────────────────────────────────────────────────────────
# Resolve requirement status from linked tests
# ──────────────────────────────────────────────────────────────────────────────
def req_status(test_names):
    """PASS if all mapped tests pass, SKIP if any skip, FAIL if any fail, MISSING if no tests found."""
    found = [test_index[n] for n in test_names if n in test_index]
    if not found:
        return "MISSING"
    statuses = {t["status"] for t in found}
    if "FAIL"  in statuses: return "FAIL"
    if "SKIP"  in statuses: return "SKIP"
    return "PASS"

STATUS_COLOR = {"PASS": C_PASS, "FAIL": C_FAIL, "SKIP": C_SKIP,
                "MISSING": "E2EFDA", "N/A": "FFFFFF"}

# ──────────────────────────────────────────────────────────────────────────────
# Workbook
# ──────────────────────────────────────────────────────────────────────────────
wb = Workbook()

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 1 – SUMMARY
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Summary"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 32
ws.column_dimensions["B"].width = 14
ws.column_dimensions["C"].width = 14
ws.column_dimensions["D"].width = 14
ws.column_dimensions["E"].width = 14
ws.column_dimensions["F"].width = 36

def hdr_cell(ws, row, col, text, bg=C_HEAD1, fg="FFFFFF", sz=11, bold=True, colspan=1):
    c = ws.cell(row, col, text)
    c.fill = fill(bg)
    c.font = font(bold=bold, color=fg, size=sz)
    c.alignment = center()
    c.border = thin_border()
    return c

def data_cell(ws, row, col, text, bg=C_WHITE, bold=False, align="left"):
    c = ws.cell(row, col, text)
    c.fill = fill(bg)
    c.font = font(bold=bold)
    c.alignment = left() if align == "left" else center()
    c.border = thin_border()
    return c

# Title
ws.merge_cells("A1:F1")
t = ws.cell(1,1,"PKCS#11 v2.40 — Test Suite Requirements Coverage Report")
t.fill = fill(C_HEAD1); t.font = font(bold=True, color="FFFFFF", size=14)
t.alignment = center(); t.border = thick_border()
ws.row_dimensions[1].height = 30

ws.merge_cells("A2:F2")
ws.cell(2,1,"Target: SoftHSM2 2.6.1 | Standard: PKCS#11 v2.40 (Cryptoki)")
ws.cell(2,1).fill = fill(C_HEAD3)
ws.cell(2,1).font = font(italic=True, size=10)
ws.cell(2,1).alignment = center()

# ── Test Results stats ───────────────────────────────────────────────────────
pass_c  = sum(1 for t in tests if t["status"]=="PASS")
fail_c  = sum(1 for t in tests if t["status"]=="FAIL")
skip_c  = sum(1 for t in tests if t["status"]=="SKIP")
total_c = len(tests)

req_pass = sum(1 for r in REQUIREMENTS if req_status(r[4])=="PASS")
req_fail = sum(1 for r in REQUIREMENTS if req_status(r[4])=="FAIL")
req_skip = sum(1 for r in REQUIREMENTS if req_status(r[4])=="SKIP")
req_miss = sum(1 for r in REQUIREMENTS if req_status(r[4])=="MISSING")
req_total = len(REQUIREMENTS)

ws.row_dimensions[4].height = 18
hdr_cell(ws, 4, 1, "Category",       bg=C_HEAD2)
hdr_cell(ws, 4, 2, "PASS",           bg=C_HEAD2)
hdr_cell(ws, 4, 3, "FAIL",           bg=C_HEAD2)
hdr_cell(ws, 4, 4, "SKIP",           bg=C_HEAD2)
hdr_cell(ws, 4, 5, "TOTAL",          bg=C_HEAD2)
hdr_cell(ws, 4, 6, "Coverage %",     bg=C_HEAD2)

for r, (cat, p, f, s, tot) in enumerate([
    ("Test Cases",    pass_c,  fail_c,  skip_c,  total_c),
    ("Requirements",  req_pass,req_fail,req_skip, req_total),
], start=5):
    ws.row_dimensions[r].height = 18
    data_cell(ws, r, 1, cat, bg=C_ALT if r%2==0 else C_WHITE, bold=True)
    data_cell(ws, r, 2, p,  bg=C_PASS, align="center")
    data_cell(ws, r, 3, f,  bg=(C_FAIL if f>0 else C_WHITE), align="center")
    data_cell(ws, r, 4, s,  bg=(C_SKIP if s>0 else C_WHITE), align="center")
    data_cell(ws, r, 5, tot, align="center")
    pct = f"{100*(p)/(tot-s) if (tot-s)>0 else 0:.1f}%"
    data_cell(ws, r, 6, pct, bg=C_PASS if f==0 else C_SKIP, align="center")

# ── Per-group breakdown ──────────────────────────────────────────────────────
ws.row_dimensions[8].height = 18
hdr_cell(ws, 8, 1, "Test Group",     bg=C_HEAD2)
hdr_cell(ws, 8, 2, "PASS",           bg=C_HEAD2)
hdr_cell(ws, 8, 3, "FAIL",           bg=C_HEAD2)
hdr_cell(ws, 8, 4, "SKIP",           bg=C_HEAD2)
hdr_cell(ws, 8, 5, "TOTAL",          bg=C_HEAD2)
hdr_cell(ws, 8, 6, "Group Description", bg=C_HEAD2)

GROUP_DESC = {
    "init":          "Library initialisation & persistent state",
    "slot_token":    "Slot and token management",
    "session":       "Session lifecycle",
    "login":         "Login / logout & PIN management",
    "objects":       "Object CRUD and search",
    "crypto_sym":    "Symmetric encryption (AES)",
    "crypto_asym":   "Asymmetric crypto (RSA, EC)",
    "digest":        "Message digest (SHA family)",
    "random":        "Random number generation",
    "key_generation":"Key generation (AES, RSA, EC, DES3)",
    "wrap_unwrap":   "Key wrap and unwrap",
    "derive":        "Key derivation (ECDH, KDF)",
    "dual_function": "Dual-function operations",
    "parallel":      "Multi-threaded session safety",
    "edge_cases":    "Edge cases and error paths",
    "req_general":   "REQ: General & slot/token (§11.4-11.5)",
    "req_session":   "REQ: Session & login (§11.6)",
    "req_objects":   "REQ: Object management (§11.7)",
    "req_crypto":    "REQ: Cryptographic ops (§11.8-11.15)",
}
groups = {}
for t in tests:
    g = t["group"]
    if g not in groups: groups[g] = {"PASS":0,"FAIL":0,"SKIP":0}
    groups[g][t["status"]] += 1

for row, (g, cnts) in enumerate(sorted(groups.items()), start=9):
    ws.row_dimensions[row].height = 16
    tot = sum(cnts.values())
    bg = C_ALT if row%2==0 else C_WHITE
    data_cell(ws, row, 1, g, bg=bg, bold=True)
    data_cell(ws, row, 2, cnts["PASS"], bg=(C_PASS if cnts["PASS"]>0 else C_WHITE), align="center")
    data_cell(ws, row, 3, cnts["FAIL"], bg=(C_FAIL if cnts["FAIL"]>0 else C_WHITE), align="center")
    data_cell(ws, row, 4, cnts["SKIP"], bg=(C_SKIP if cnts["SKIP"]>0 else C_WHITE), align="center")
    data_cell(ws, row, 5, tot, align="center")
    data_cell(ws, row, 6, GROUP_DESC.get(g,""), bg=bg)

# ── Legend ───────────────────────────────────────────────────────────────────
leg_row = 9 + len(groups) + 2
for col, (label, color) in enumerate([
    ("PASS – test executed and succeeded", C_PASS),
    ("FAIL – test executed and failed", C_FAIL),
    ("SKIP – mechanism not supported by SoftHSM2 2.6.1", C_SKIP),
], start=1):
    c = ws.cell(leg_row, col, label)
    c.fill = fill(color); c.font = font(size=9); c.alignment = left(); c.border = thin_border()


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 2 – REQUIREMENTS
# ══════════════════════════════════════════════════════════════════════════════
wr = wb.create_sheet("Requirements")
wr.sheet_view.showGridLines = False
wr.freeze_panes = "A3"

cols = [("Req ID",14),("Section",10),("Function",22),("Requirement Description",52),
        ("Mapped Tests",48),("Status",10),("Notes",28)]
for i,(h,w) in enumerate(cols,1):
    wr.column_dimensions[get_column_letter(i)].width = w

# Title row
wr.merge_cells(f"A1:{get_column_letter(len(cols))}1")
tc = wr.cell(1,1,"PKCS#11 v2.40 — Requirements Traceability")
tc.fill = fill(C_HEAD1); tc.font = font(bold=True,color="FFFFFF",size=13)
tc.alignment = center(); tc.border = thick_border()
wr.row_dimensions[1].height = 26

# Header row
for i,(h,_) in enumerate(cols,1):
    hdr_cell(wr, 2, i, h, bg=C_HEAD2)
wr.row_dimensions[2].height = 20

section_colors = {
    "§11.4":"BDD7EE","§11.5":"9DC3E6","§11.6":"2E74B5",
    "§11.7":"1F4E79","§11.8":"70AD47","§11.9":"A9D18E",
    "§11.10":"FFC000","§11.11":"FFD966","§11.12":"F4B183",
    "§11.13":"C9C9C9","§11.14":"7030A0","§11.15":"E2EFDA",
    "§11.x":"BFBFBF",
}

prev_section = None
for row, (req_id, section, func, desc, tnames) in enumerate(REQUIREMENTS, start=3):
    status = req_status(tnames)
    bg_row = C_ALT if row%2==0 else C_WHITE
    sec_bg = section_colors.get(section, C_WHITE)

    wr.row_dimensions[row].height = 30

    c = wr.cell(row, 1, req_id)
    c.fill = fill(bg_row); c.font = font(bold=True, size=9)
    c.alignment = center(); c.border = thin_border()

    c = wr.cell(row, 2, section)
    c.fill = fill(sec_bg); c.font = font(size=9, color="FFFFFF" if section in ("§11.6","§11.7","§11.14") else "000000")
    c.alignment = center(); c.border = thin_border()

    c = wr.cell(row, 3, func)
    c.fill = fill(bg_row); c.font = font(bold=True, size=9)
    c.alignment = left(); c.border = thin_border()

    c = wr.cell(row, 4, desc)
    c.fill = fill(bg_row); c.font = font(size=9)
    c.alignment = left(); c.border = thin_border()

    mapped = "\n".join(tnames)
    c = wr.cell(row, 5, mapped)
    c.fill = fill(bg_row); c.font = font(size=8, italic=True)
    c.alignment = left(); c.border = thin_border()

    c = wr.cell(row, 6, status)
    c.fill = fill(STATUS_COLOR.get(status, C_WHITE))
    c.font = font(bold=True, size=9)
    c.alignment = center(); c.border = thin_border()

    notes = ""
    if status == "SKIP":
        notes = "Mechanism not supported by SoftHSM2 2.6.1"
    elif status == "MISSING":
        notes = "No test mapped yet"
    c = wr.cell(row, 7, notes)
    c.fill = fill(bg_row); c.font = font(size=8, italic=True, color="595959")
    c.alignment = left(); c.border = thin_border()


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 3 – TEST RESULTS
# ══════════════════════════════════════════════════════════════════════════════
wt = wb.create_sheet("Test Results")
wt.sheet_view.showGridLines = False
wt.freeze_panes = "A3"

tcols = [("Test Group",20),("Test Name",52),("Status",10),
         ("Req IDs Covered",32),("Notes",28)]
for i,(h,w) in enumerate(tcols,1):
    wt.column_dimensions[get_column_letter(i)].width = w

wt.merge_cells(f"A1:{get_column_letter(len(tcols))}1")
tc = wt.cell(1,1,"PKCS#11 Test Suite — Full Test Results")
tc.fill = fill(C_HEAD1); tc.font = font(bold=True,color="FFFFFF",size=13)
tc.alignment = center(); tc.border = thick_border()
wt.row_dimensions[1].height = 26

for i,(h,_) in enumerate(tcols,1):
    hdr_cell(wt, 2, i, h, bg=C_HEAD2)
wt.row_dimensions[2].height = 20

# Build reverse map: test_name → [req_ids]
test_to_reqs = {}
for req_id, section, func, desc, tnames in REQUIREMENTS:
    for tn in tnames:
        test_to_reqs.setdefault(tn, []).append(req_id)

prev_group = None
for row, t in enumerate(tests, start=3):
    wt.row_dimensions[row].height = 16
    bg = C_ALT if t["group"] != prev_group else (C_WHITE if row%2==0 else "F5F9FF")
    prev_group = t["group"]

    c = wt.cell(row, 1, t["group"])
    c.fill = fill(C_HEAD3 if bg==C_ALT else bg)
    c.font = font(bold=True, size=9); c.alignment = left(); c.border = thin_border()

    c = wt.cell(row, 2, t["name"])
    c.fill = fill(bg); c.font = font(size=9); c.alignment = left(); c.border = thin_border()

    status = t["status"]
    c = wt.cell(row, 3, status)
    c.fill = fill(STATUS_COLOR.get(status, C_WHITE))
    c.font = font(bold=True, size=9); c.alignment = center(); c.border = thin_border()

    reqs = test_to_reqs.get(t["name"], [])
    c = wt.cell(row, 4, ", ".join(reqs) if reqs else "—")
    c.fill = fill(bg); c.font = font(size=8, italic=True); c.alignment = left(); c.border = thin_border()

    notes = ""
    if status == "SKIP":
        notes = "Mechanism not supported by target"
    c = wt.cell(row, 5, notes)
    c.fill = fill(bg); c.font = font(size=8, italic=True, color="595959")
    c.alignment = left(); c.border = thin_border()


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 4 – COVERAGE MATRIX  (requirement section × test group)
# ══════════════════════════════════════════════════════════════════════════════
wm = wb.create_sheet("Coverage Matrix")
wm.sheet_view.showGridLines = False

SECTIONS = ["§11.4","§11.5","§11.6","§11.7","§11.8","§11.9",
            "§11.10","§11.11","§11.12","§11.13","§11.14","§11.15","§11.x"]

SECTION_NAMES = {
    "§11.4": "General (Init/Finalize/GetInfo)",
    "§11.5": "Slot & Token Management",
    "§11.6": "Session & Login",
    "§11.7": "Object Management",
    "§11.8": "Encryption",
    "§11.9": "Decryption",
    "§11.10":"Message Digest",
    "§11.11":"Sign / MAC",
    "§11.12":"Verify",
    "§11.13":"Dual-Function",
    "§11.14":"Key Management",
    "§11.15":"Random",
    "§11.x": "Parallel / Other",
}

GROUP_ORDER = list(GROUP_DESC.keys())

# Build matrix: section → group → list of statuses
matrix = {s: {g: [] for g in GROUP_ORDER} for s in SECTIONS}
for req_id, section, func, desc, tnames in REQUIREMENTS:
    status = req_status(tnames)
    for tn in tnames:
        if tn in test_index:
            g = test_index[tn]["group"]
            if g in matrix[section]:
                matrix[section][g].append(status)

def cell_val(statuses):
    if not statuses: return ""
    if "FAIL" in statuses: return "FAIL"
    if "PASS" in statuses: return "✓"
    if "SKIP" in statuses: return "SKIP"
    return ""

wm.merge_cells(f"A1:{get_column_letter(len(GROUP_ORDER)+2)}1")
tc = wm.cell(1,1,"Coverage Matrix — PKCS#11 Sections × Test Groups")
tc.fill = fill(C_HEAD1); tc.font = font(bold=True,color="FFFFFF",size=13)
tc.alignment = center(); tc.border = thick_border()
wm.row_dimensions[1].height = 26

# Column headers (group names)
wm.column_dimensions["A"].width = 10
wm.column_dimensions["B"].width = 34
for col, g in enumerate(GROUP_ORDER, start=3):
    wm.column_dimensions[get_column_letter(col)].width = 12
    c = wm.cell(2, col, g)
    c.fill = fill(C_HEAD2); c.font = font(bold=True,color="FFFFFF",size=8)
    c.alignment = Alignment(horizontal="center",vertical="center",
                             wrap_text=True, text_rotation=45)
    c.border = thin_border()
wm.row_dimensions[2].height = 72

hdr_cell(wm, 2, 1, "Section",  bg=C_HEAD2, sz=9)
hdr_cell(wm, 2, 2, "Description", bg=C_HEAD2, sz=9)

for row, sec in enumerate(SECTIONS, start=3):
    wm.row_dimensions[row].height = 22
    sec_bg = section_colors.get(sec, C_WHITE)
    fg = "FFFFFF" if sec in ("§11.6","§11.7","§11.14") else "000000"

    c = wm.cell(row, 1, sec)
    c.fill = fill(sec_bg); c.font = font(bold=True,size=9,color=fg)
    c.alignment = center(); c.border = thin_border()

    c = wm.cell(row, 2, SECTION_NAMES.get(sec,""))
    c.fill = fill(sec_bg); c.font = font(size=9,color=fg)
    c.alignment = left(); c.border = thin_border()

    for col, g in enumerate(GROUP_ORDER, start=3):
        statuses = matrix[sec][g]
        val = cell_val(statuses)
        cell_bg = C_WHITE
        if val == "✓":   cell_bg = C_PASS
        elif val == "FAIL": cell_bg = C_FAIL
        elif val == "SKIP": cell_bg = C_SKIP

        c = wm.cell(row, col, val)
        c.fill = fill(cell_bg)
        c.font = font(bold=(val!=""), size=9,
                      color="006100" if val=="✓" else
                            "9C0006" if val=="FAIL" else
                            "7D6608" if val=="SKIP" else "000000")
        c.alignment = center(); c.border = thin_border()

# Legend
leg = len(SECTIONS) + 4
for col, (txt, clr) in enumerate([
    ("✓  PASS – at least one mapped test passed", C_PASS),
    ("FAIL – at least one mapped test failed",    C_FAIL),
    ("SKIP – mechanism not supported",            C_SKIP),
    ("(blank) – no tests in this group for section","FFFFFF"),
], start=1):
    c = wm.cell(leg, col, txt)
    c.fill = fill(clr); c.font = font(size=9); c.alignment = left(); c.border = thin_border()

# ──────────────────────────────────────────────────────────────────────────────
# Tab colours
# ──────────────────────────────────────────────────────────────────────────────
wb["Summary"].sheet_properties.tabColor       = "1F4E79"
wb["Requirements"].sheet_properties.tabColor  = "2E74B5"
wb["Test Results"].sheet_properties.tabColor  = "70AD47"
wb["Coverage Matrix"].sheet_properties.tabColor = "7030A0"

out = "/home/user/Claude/PKCS11_Requirements_Coverage.xlsx"
wb.save(out)
print(f"\nSaved: {out}")
print(f"Test cases: {total_c} (PASS={pass_c}, FAIL={fail_c}, SKIP={skip_c})")
print(f"Requirements: {req_total} (PASS={req_pass}, FAIL={req_fail}, SKIP={req_skip}, MISSING={req_miss})")

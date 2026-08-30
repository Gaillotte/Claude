#!/usr/bin/env python3
"""
Génère un rapport Excel (.xlsx) à partir du rapport JSON du pipeline AI Transit.

Onglet 0 — Résumé  : lien du repo, date du scan, hash global du repo
Onglet 1 — Fichiers : une ligne par fichier scanné avec statut et message
Onglet 2 — Findings : FAIL + WARN uniquement, triés par sévérité
"""
from __future__ import annotations

import json
import sys
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side
    )
except ImportError:
    print("[ERREUR] openpyxl manquant — installez-le avec : pip install openpyxl",
          file=sys.stderr)
    sys.exit(1)

# ── Couleurs ─────────────────────────────────────────────────────────────────
COLOR_HEADER   = "1F3864"   # bleu marine
COLOR_PASS     = "C6EFCE"   # vert clair
COLOR_FAIL     = "FFC7CE"   # rouge clair
COLOR_WARN     = "FFEB9C"   # jaune clair
COLOR_TITLE_BG = "2E4057"
FONT_WHITE     = Font(color="FFFFFF", bold=True, name="Calibri")
FONT_BOLD      = Font(bold=True, name="Calibri")
FONT_NORMAL    = Font(name="Calibri")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _set_col_width(ws, col_letter: str, width: float) -> None:
    ws.column_dimensions[col_letter].width = width


def _header_cell(ws, row: int, col: int, value: str) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = FONT_WHITE
    cell.fill = _fill(COLOR_HEADER)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = THIN_BORDER


def _data_cell(ws, row: int, col: int, value: str,
               fill_color: str | None = None) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = FONT_NORMAL
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = THIN_BORDER
    if fill_color:
        cell.fill = _fill(fill_color)


def _rel(path: str, scan_dir: str) -> str:
    """Display path relative to the scanned directory.

    Report paths are absolute and point inside WORK_DIR/fetch/repo_<ts>/.
    Showing that to a reader leaks the internal layout and buries the part
    that matters, so strip the scan-directory prefix for display.
    """
    if scan_dir and path.startswith(scan_dir.rstrip("/") + "/"):
        return path[len(scan_dir.rstrip("/")) + 1:]
    return path


def _verdict_fill(status: str) -> str | None:
    return {
        "PASS": COLOR_PASS,
        "FAIL": COLOR_FAIL,
        "WARN": COLOR_WARN,
    }.get(status.upper())


# ── Onglet 0 : Résumé ────────────────────────────────────────────────────────
def build_sheet_summary(wb: Workbook, data: dict) -> None:
    ws = wb.active
    ws.title = "Résumé"

    title_fill  = _fill(COLOR_TITLE_BG)
    title_font  = Font(color="FFFFFF", bold=True, size=14, name="Calibri")

    ws.merge_cells("A1:B1")
    title = ws["A1"]
    title.value = "AI Transit Pipeline — Rapport de scan"
    title.font  = title_font
    title.fill  = title_fill
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    rows = [
        ("Dépôt / Source",  data.get("repo_input") or data.get("directory", "—")),
        ("Date du scan",    data.get("timestamp", "—")),
        ("Hash global SHA-256", data.get("repo_hash", "—")),
        ("Verdict",         data.get("verdict", "—")),
        ("Fichiers PASS",   data.get("summary", {}).get("pass", 0)),
        ("Fichiers WARN",   data.get("summary", {}).get("warn", 0)),
        ("Fichiers FAIL",   data.get("summary", {}).get("fail", 0)),
        ("Répertoire scanné", data.get("directory", "—")),
    ]

    for i, (label, value) in enumerate(rows, start=2):
        label_cell = ws.cell(row=i, column=1, value=label)
        label_cell.font  = FONT_BOLD
        label_cell.fill  = _fill("D9E1F2")
        label_cell.border = THIN_BORDER
        label_cell.alignment = Alignment(vertical="center")

        value_cell = ws.cell(row=i, column=2, value=str(value))
        value_cell.font   = FONT_NORMAL
        value_cell.border = THIN_BORDER
        value_cell.alignment = Alignment(vertical="center", wrap_text=True)

        if label == "Verdict":
            value_cell.fill = _fill(
                COLOR_PASS if value == "PASS"
                else COLOR_WARN if value == "WARN"
                else COLOR_FAIL
            )
            value_cell.font = Font(bold=True, name="Calibri")

        ws.row_dimensions[i].height = 20

    _set_col_width(ws, "A", 28)
    _set_col_width(ws, "B", 80)


# ── Onglet 1 : Fichiers ───────────────────────────────────────────────────────
def build_sheet_files(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet(title="Fichiers")

    headers = ["#", "Fichier", "Type", "Statut", "Message / Finding"]
    for col, h in enumerate(headers, start=1):
        _header_cell(ws, 1, col, h)
    ws.row_dimensions[1].height = 22

    file_results: dict = data.get("file_results", {})
    scan_dir: str = data.get("directory", "")

    # Tri : FAIL en premier, puis WARN, puis PASS
    order = {"FAIL": 0, "WARN": 1, "PASS": 2}
    sorted_files = sorted(
        file_results.items(),
        key=lambda kv: order.get(kv[1].get("status", "PASS").upper(), 3)
    )

    for idx, (filepath, info) in enumerate(sorted_files, start=1):
        row = idx + 1
        status  = info.get("status", "?").upper()
        message = info.get("message", "").rstrip(" | ")
        suffix  = Path(filepath).suffix.lower() or Path(filepath).name
        fill    = _verdict_fill(status)

        _data_cell(ws, row, 1, idx)
        _data_cell(ws, row, 2, _rel(filepath, scan_dir), fill)
        _data_cell(ws, row, 3, suffix, fill)
        _data_cell(ws, row, 4, status, fill)
        _data_cell(ws, row, 5, message, fill)

        ws.row_dimensions[row].height = 18

    _set_col_width(ws, "A", 6)
    _set_col_width(ws, "B", 70)
    _set_col_width(ws, "C", 12)
    _set_col_width(ws, "D", 10)
    _set_col_width(ws, "E", 60)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{len(file_results) + 1}"


# ── Onglet 2 : Findings (FAIL + WARN only) ───────────────────────────────────
def build_sheet_findings(wb: "Workbook", data: dict) -> None:
    ws = wb.create_sheet(title="Findings")

    headers = ["#", "Severity", "File", "Finding / Rule", "Status"]
    for col, h in enumerate(headers, start=1):
        _header_cell(ws, 1, col, h)
    ws.row_dimensions[1].height = 22

    file_results: dict = data.get("file_results", {})
    scan_dir: str = data.get("directory", "")

    rows = []
    for filepath, info in file_results.items():
        status = info.get("status", "PASS").upper()
        if status not in ("FAIL", "WARN"):
            continue
        message = info.get("message", "").rstrip(" | ")
        # Split on " | " (the delimiter used by record_fail accumulation).
        # A bare "|" split would incorrectly break on pipe chars in finding text.
        findings = [m.strip() for m in message.split(" | ") if m.strip()]
        if not findings:
            findings = ["(no details)"]
        for finding in findings:
            # Each message carries its own [WARN]/[FAIL] tag from the scanner.
            # Severity must come from the message itself, never from the file's
            # overall status — otherwise a missing-tool WARN recorded against a
            # file that also has a real FAIL would be reported as HIGH.
            if finding.startswith("[FAIL]"):
                entry_status = "FAIL"
                finding = finding[len("[FAIL]"):].strip()
            elif finding.startswith("[WARN]"):
                entry_status = "WARN"
                finding = finding[len("[WARN]"):].strip()
            else:
                # Untagged (older report format): fall back to file status.
                entry_status = status

            f_upper = finding.upper()
            if entry_status == "FAIL":
                if ":CRITICAL:" in f_upper:
                    sev, sev_order = "CRITICAL", 0
                else:
                    sev, sev_order = "HIGH", 1
            else:  # WARN
                if ":MEDIUM:" in f_upper:
                    sev, sev_order = "MEDIUM", 2
                else:
                    sev, sev_order = "LOW", 3
            rows.append((sev_order, sev, filepath, finding, entry_status))

    # Sort: CRITICAL first, then HIGH, MEDIUM, LOW
    rows.sort(key=lambda r: r[0])

    status_fill = {"FAIL": COLOR_FAIL, "WARN": COLOR_WARN}
    sev_fill = {"CRITICAL": "FF0000", "HIGH": COLOR_FAIL,
                "MEDIUM": COLOR_WARN, "LOW": "D9D9D9"}

    for idx, (_, sev, filepath, finding, status) in enumerate(rows, start=1):
        row = idx + 1
        fill = status_fill.get(status, None)
        _data_cell(ws, row, 1, idx)
        cell = ws.cell(row=row, column=2, value=sev)
        cell.font = Font(bold=True, name="Calibri",
                         color="FFFFFF" if sev in ("CRITICAL", "HIGH") else "000000")
        cell.fill = _fill(sev_fill.get(sev, "FFFFFF"))
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        _data_cell(ws, row, 3, _rel(filepath, scan_dir), fill)
        _data_cell(ws, row, 4, finding, fill)
        _data_cell(ws, row, 5, status, fill)
        ws.row_dimensions[row].height = 18

    _set_col_width(ws, "A", 6)
    _set_col_width(ws, "B", 12)
    _set_col_width(ws, "C", 60)
    _set_col_width(ws, "D", 80)
    _set_col_width(ws, "E", 10)

    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = f"A1:E{len(rows) + 1}"


# ── Point d'entrée ────────────────────────────────────────────────────────────
def main() -> None:
    if len(sys.argv) < 3:
        print(f"Usage : {sys.argv[0]} <rapport.json> <sortie.xlsx>", file=sys.stderr)
        sys.exit(1)

    json_path  = Path(sys.argv[1])
    xlsx_path  = Path(sys.argv[2])

    if not json_path.exists():
        print(f"[ERREUR] Fichier JSON introuvable : {json_path}", file=sys.stderr)
        sys.exit(1)

    with json_path.open(encoding="utf-8") as fh:
        data = json.load(fh)

    wb = Workbook()
    build_sheet_summary(wb, data)
    build_sheet_files(wb, data)
    build_sheet_findings(wb, data)

    wb.save(xlsx_path)
    print(f"[OK] Rapport Excel généré : {xlsx_path}")


if __name__ == "__main__":
    main()

"""Unit tests for the Excel performance report generator (no HSM required)."""

import pytest
from pathlib import Path

from pkcs11_app.report import PerfEntry, PerfReport, _fmt_bytes


pytestmark = pytest.mark.unit


# ── _fmt_bytes ────────────────────────────────────────────────────────────────

@pytest.mark.parametrize("n,expected", [
    (0,       "—"),
    (1,       "1 B"),
    (512,     "512 B"),
    (1023,    "1023 B"),
    (1024,    "1 KB"),
    (4096,    "4 KB"),
    (65536,   "64 KB"),
    (1048576, "1 MB"),
    (2097152, "2 MB"),
])
def test_fmt_bytes(n, expected):
    assert _fmt_bytes(n) == expected


# ── PerfEntry ─────────────────────────────────────────────────────────────────

def test_perf_entry_defaults():
    e = PerfEntry(
        category="AES", mechanism="AES_CBC_PAD", operation="encrypt",
        key_bits=256, payload_bytes=1024,
        ops_per_sec=50000.0, latency_ms=0.02, throughput_mbs=47.7,
    )
    assert e.notes == ""
    assert e.category == "AES"
    assert e.throughput_mbs == pytest.approx(47.7)


# ── PerfReport.write ──────────────────────────────────────────────────────────

def _make_report() -> PerfReport:
    """Build a PerfReport covering multiple categories for full sheet coverage."""
    r = PerfReport()
    categories = {
        "AES":     [("AES_CBC_PAD",  "encrypt", 256, 1024,    50000.0, 0.02,  47.7),
                    ("AES_GCM",      "encrypt", 128, 65536,   45000.0, 0.022, 42.0),
                    ("AES_ECB",      "encrypt", 128, 1024,    76000.0, 0.013, 72.5),
                    ("AES_CTR",      "encrypt", 256, 65536,   60000.0, 0.017, 57.2)],
        "DES3":    [("DES3_CBC_PAD", "encrypt", 168, 1024,    30000.0, 0.033, 28.6)],
        "RSA":     [("SHA256_RSA_PKCS", "sign",    2048, 0,  1200.0, 0.83,  0.0),
                    ("SHA256_RSA_PKCS", "verify",  2048, 0, 18000.0, 0.056, 0.0),
                    ("RSA_PKCS_OAEP",  "encrypt",  2048, 64, 9000.0, 0.11,  0.0),
                    ("RSA_PKCS_OAEP",  "decrypt",  2048, 64,  500.0, 2.0,   0.0)],
        "EC":      [("ECDSA",        "sign",    256,  0,  2000.0, 0.50, 0.0),
                    ("ECDSA",        "verify",  256,  0, 10000.0, 0.10, 0.0),
                    ("ECDH1_DERIVE", "derive",  256,  0,  1500.0, 0.67, 0.0)],
        "EdDSA":   [("EDDSA",        "sign",    256, 64,  2500.0, 0.40, 0.0),
                    ("EDDSA",        "verify",  256, 64,  8000.0, 0.125,0.0)],
        "DSA":     [("DSA_SHA256",   "sign",    1024, 0,  1000.0, 1.0,  0.0),
                    ("DSA_SHA256",   "verify",  1024, 0,  5000.0, 0.2,  0.0)],
        "DH":      [("DH_PKCS_DERIVE","derive", 1024, 0,   800.0, 1.25, 0.0)],
        "HMAC":    [("HMAC_SHA256",  "sign",    256, 4096, 55000.0, 0.018, 215.0),
                    ("AES_CMAC",     "sign",    128, 1024, 72000.0, 0.014, 70.3)],
        "Digest":  [("SHA256",       "digest",  256, 65536, 80000.0, 0.013, 5000.0)],
        "KeyGen":  [("AES_KEY_GEN",  "keygen",  256, 0, 5000.0, 0.20, 0.0),
                    ("RSA_PKCS_KEY_PAIR_GEN", "keygen", 2048, 0, 5.0, 200.0, 0.0)],
        "KeyWrap": [("AES_KEY_WRAP", "wrap",    256, 0, 40000.0, 0.025, 0.0)],
    }
    for cat, rows in categories.items():
        for mech, op, kb, pb, ops, lat, thr in rows:
            r.add(PerfEntry(
                category=cat, mechanism=mech, operation=op,
                key_bits=kb, payload_bytes=pb,
                ops_per_sec=ops, latency_ms=lat, throughput_mbs=thr,
            ))
    return r


def test_write_creates_file(tmp_path):
    report = _make_report()
    out = tmp_path / "test_report.xlsx"
    result = report.write(out)
    assert result == out
    assert out.exists()
    assert out.stat().st_size > 5000


def test_workbook_sheets(tmp_path):
    import openpyxl
    report = _make_report()
    out = tmp_path / "sheets.xlsx"
    report.write(out)
    wb = openpyxl.load_workbook(out)
    assert "Résumé" in wb.sheetnames
    assert "Vue d'ensemble" in wb.sheetnames
    # Every category becomes its own sheet
    for cat in ("AES", "DES3", "RSA", "EC", "EdDSA", "DSA", "DH",
                "HMAC", "Digest", "KeyGen", "KeyWrap"):
        assert cat in wb.sheetnames, f"Missing sheet: {cat}"


def test_summary_row_count(tmp_path):
    import openpyxl
    report = _make_report()
    out = tmp_path / "rows.xlsx"
    report.write(out)
    wb = openpyxl.load_workbook(out)
    ws = wb["Résumé"]
    # row 1=title, 2=stats, 3=header → data starts at 4
    data_rows = ws.max_row - 3
    assert data_rows == len(report.entries)


def test_summary_numeric_values(tmp_path):
    import openpyxl
    report = _make_report()
    out = tmp_path / "vals.xlsx"
    report.write(out)
    wb = openpyxl.load_workbook(out)
    ws = wb["Résumé"]
    # Check first data row (row 4): ops_per_sec in column 6
    ops_val = ws.cell(row=4, column=6).value
    assert isinstance(ops_val, (int, float))
    assert ops_val > 0


def test_empty_report_no_file(tmp_path):
    """An empty PerfReport should still write a valid (header-only) workbook."""
    report = PerfReport()
    out = tmp_path / "empty.xlsx"
    report.write(out)
    assert out.exists()


def test_overview_chart_sheet(tmp_path):
    import openpyxl
    report = _make_report()
    out = tmp_path / "chart.xlsx"
    report.write(out)
    wb = openpyxl.load_workbook(out)
    ws = wb["Vue d'ensemble"]
    # Should have at least one chart
    assert len(ws._charts) >= 1


def test_category_sheet_has_chart(tmp_path):
    import openpyxl
    report = _make_report()
    out = tmp_path / "catchart.xlsx"
    report.write(out)
    wb = openpyxl.load_workbook(out)
    ws = wb["AES"]
    assert len(ws._charts) >= 1


def test_zero_payload_shown_as_dash(tmp_path):
    import openpyxl
    report = PerfReport()
    report.add(PerfEntry("RSA", "RSA_PKCS", "sign", 2048, 0, 1200.0, 0.83))
    out = tmp_path / "nodash.xlsx"
    report.write(out)
    wb = openpyxl.load_workbook(out)
    ws = wb["Résumé"]
    # Column 5 = payload
    payload_cell = ws.cell(row=4, column=5).value
    assert payload_cell == "—"


def test_throughput_zero_shown_as_dash(tmp_path):
    import openpyxl
    report = PerfReport()
    report.add(PerfEntry("RSA", "RSA_PKCS", "sign", 2048, 0, 1200.0, 0.83, throughput_mbs=0.0))
    out = tmp_path / "nothr.xlsx"
    report.write(out)
    wb = openpyxl.load_workbook(out)
    ws = wb["Résumé"]
    thr_cell = ws.cell(row=4, column=8).value
    assert thr_cell == "—"


def test_overview_best_ops_per_mechanism(tmp_path):
    """When the same mechanism+operation appears multiple times, only the max ops/s is shown."""
    import openpyxl
    report = PerfReport()
    # Same mechanism, two entries with different ops — overview should keep the higher one
    report.add(PerfEntry("AES", "AES_CBC_PAD", "encrypt", 256, 1024, 30000.0, 0.033, 28.6))
    report.add(PerfEntry("AES", "AES_CBC_PAD", "encrypt", 128, 65536, 75000.0, 0.013, 70.0))
    out = tmp_path / "best.xlsx"
    report.write(out)
    wb = openpyxl.load_workbook(out)
    ws = wb["Vue d'ensemble"]
    # Find the ops/s value for the AES_CBC_PAD encrypt row
    values = [ws.cell(r, 2).value for r in range(3, ws.max_row + 1)
              if ws.cell(r, 2).value is not None]
    assert 75000.0 in values
    assert 30000.0 not in values


def test_add_bar_chart_empty_entries(tmp_path):
    """_add_bar_chart should silently return when given an empty list."""
    import openpyxl
    from pkcs11_app.report import PerfReport
    report = PerfReport()
    wb = openpyxl.Workbook()
    ws = wb.active
    # Direct call with empty list — should not raise
    report._add_bar_chart(ws, [], start_row=1)
    assert len(ws._charts) == 0


def test_overview_duplicate_key_lower_ops_ignored(tmp_path):
    """Second entry with same mechanism+op but lower ops/s must NOT overwrite the best."""
    import openpyxl
    report = PerfReport()
    report.add(PerfEntry("AES", "AES_CBC_PAD", "encrypt", 256, 65536, 75000.0, 0.013, 70.0))
    # Lower entry added after — the max should stay 75000
    report.add(PerfEntry("AES", "AES_CBC_PAD", "encrypt", 256, 1024, 20000.0, 0.050, 19.0))
    out = tmp_path / "dup.xlsx"
    report.write(out)
    wb = openpyxl.load_workbook(out)
    ws = wb["Vue d'ensemble"]
    values = [ws.cell(r, 2).value for r in range(3, ws.max_row + 1)
              if ws.cell(r, 2).value is not None]
    assert 75000.0 in values
    assert 20000.0 not in values

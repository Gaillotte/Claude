"""Excel performance report generator for PKCS11 benchmarks."""

from __future__ import annotations

import datetime
from dataclasses import dataclass, field
from pathlib import Path
from typing import List

import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.utils import get_column_letter


# ── Data model ────────────────────────────────────────────────────────────────

@dataclass
class PerfEntry:
    category: str       # "AES", "DES3", "RSA", "EC", "EdDSA", "DSA", "HMAC", "Digest", "KeyGen", "KeyWrap"
    mechanism: str      # "AES_CBC_PAD", "RSA_PKCS", …
    operation: str      # "encrypt", "decrypt", "sign", "verify", "keygen", "derive"
    key_bits: int       # key size in bits (0 if N/A)
    payload_bytes: int  # data size (0 for keygen / derive)
    ops_per_sec: float
    latency_ms: float
    throughput_mbs: float = 0.0
    notes: str = ""


# ── Color palette per category ────────────────────────────────────────────────

_CAT_COLORS = {
    "AES":     "D6E4F0",
    "DES3":    "D5F5E3",
    "RSA":     "FAD7A0",
    "EC":      "F9EBEA",
    "EdDSA":   "FDEDEC",
    "DSA":     "F5EEF8",
    "DH":      "EBF5FB",
    "HMAC":    "EAFAF1",
    "Digest":  "FEF9E7",
    "KeyGen":  "F2F3F4",
    "KeyWrap": "E8DAEF",
}
_DEFAULT_COLOR = "FFFFFF"

_HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
_TITLE_FILL   = PatternFill("solid", fgColor="2E75B6")
_HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
_TITLE_FONT   = Font(bold=True, color="FFFFFF", name="Calibri", size=14)
_LABEL_FONT   = Font(bold=True, name="Calibri", size=10)
_DATA_FONT    = Font(name="Calibri", size=10)
_BORDER = Border(
    left=Side(style="thin", color="BDBDBD"),
    right=Side(style="thin", color="BDBDBD"),
    top=Side(style="thin", color="BDBDBD"),
    bottom=Side(style="thin", color="BDBDBD"),
)

_COLUMNS = [
    ("Catégorie",        16),
    ("Mécanisme",        26),
    ("Opération",        12),
    ("Clé (bits)",       10),
    ("Données",          12),
    ("Ops/s",            14),
    ("Latence (ms)",     14),
    ("Débit (MB/s)",     14),
    ("Notes",            22),
]

# ── Helpers ───────────────────────────────────────────────────────────────────

def _fmt_bytes(n: int) -> str:
    if n == 0:
        return "—"
    if n < 1024:
        return f"{n} B"
    if n < 1024 ** 2:
        return f"{n // 1024} KB"
    return f"{n // (1024 ** 2)} MB"


def _cell_fill(category: str) -> PatternFill:
    color = _CAT_COLORS.get(category, _DEFAULT_COLOR)
    return PatternFill("solid", fgColor=color)



def _header_row(ws, row: int, values: list, widths: list):
    for col, (val, _) in enumerate(zip(values, widths), start=1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _BORDER


# ── Report builder ─────────────────────────────────────────────────────────────

class PerfReport:
    """Accumulate PerfEntry records then write an Excel workbook."""

    def __init__(self):
        self.entries: List[PerfEntry] = []

    def add(self, entry: PerfEntry):
        self.entries.append(entry)

    # -- Main entry point -------------------------------------------------------

    def write(self, path: str | Path = "perf_report.xlsx"):
        path = Path(path)
        wb = openpyxl.Workbook()

        self._build_summary(wb)
        self._build_category_sheets(wb)
        self._build_charts_sheet(wb)
        wb.save(path)
        return path

    # -- Summary sheet ---------------------------------------------------------

    def _build_summary(self, wb: openpyxl.Workbook):
        ws = wb.active if wb.active else wb.create_sheet()
        ws.title = "Résumé"

        # Title
        ws.merge_cells("A1:I1")
        title_cell = ws["A1"]
        title_cell.value = (
            f"Rapport de performance PKCS#11 — SoftHSM2  |  "
            f"Généré le {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}"
        )
        title_cell.font = _TITLE_FONT
        title_cell.fill = _TITLE_FILL
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        # Stats band
        ws.merge_cells("A2:I2")
        stats_cell = ws["A2"]
        total = len(self.entries)
        cats = len({e.category for e in self.entries})
        mechs = len({e.mechanism for e in self.entries})
        stats_cell.value = (
            f"   {total} mesures  ·  {cats} catégories  ·  {mechs} mécanismes uniques"
        )
        stats_cell.font = Font(italic=True, color="FFFFFF", name="Calibri", size=10)
        stats_cell.fill = PatternFill("solid", fgColor="2980B9")
        stats_cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[2].height = 18

        # Column headers
        col_names = [c[0] for c in _COLUMNS]
        _header_row(ws, 3, col_names, _COLUMNS)
        ws.row_dimensions[3].height = 30

        # Set column widths
        for col, (_, w) in enumerate(_COLUMNS, start=1):
            ws.column_dimensions[get_column_letter(col)].width = w

        # Data rows
        prev_cat = None
        for row_i, e in enumerate(self.entries, start=4):
            fill = _cell_fill(e.category)
            vals = [
                e.category,
                e.mechanism,
                e.operation,
                e.key_bits if e.key_bits else "—",
                _fmt_bytes(e.payload_bytes),
                round(e.ops_per_sec, 1),
                round(e.latency_ms, 4),
                round(e.throughput_mbs, 2) if e.throughput_mbs else "—",
                e.notes,
            ]
            for col, val in enumerate(vals, start=1):
                c = ws.cell(row=row_i, column=col, value=val)
                c.fill = fill
                c.font = _DATA_FONT
                c.border = _BORDER
                c.alignment = Alignment(vertical="center")
                if col in (6, 7, 8):  # numeric cols
                    c.alignment = Alignment(horizontal="right", vertical="center")

            # Bold category name when it changes
            if e.category != prev_cat:
                ws.cell(row=row_i, column=1).font = _LABEL_FONT
                prev_cat = e.category

            ws.row_dimensions[row_i].height = 16

        # Freeze panes
        ws.freeze_panes = "A4"

        # Auto-filter
        last_row = 3 + len(self.entries)
        ws.auto_filter.ref = f"A3:I{last_row}"

    # -- Per-category sheets ---------------------------------------------------

    def _build_category_sheets(self, wb: openpyxl.Workbook):
        categories = {}
        for e in self.entries:
            categories.setdefault(e.category, []).append(e)

        for cat, entries in categories.items():
            ws = wb.create_sheet(title=cat[:31])  # max 31 chars

            # Title
            ws.merge_cells("A1:I1")
            tc = ws["A1"]
            tc.value = f"Performance — {cat}"
            tc.font = _TITLE_FONT
            tc.fill = PatternFill("solid", fgColor=_CAT_COLORS.get(cat, "2E75B6").replace("FF", "2E"))
            tc.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 26

            col_names = [c[0] for c in _COLUMNS]
            _header_row(ws, 2, col_names, _COLUMNS)
            ws.row_dimensions[2].height = 28

            for col, (_, w) in enumerate(_COLUMNS, start=1):
                ws.column_dimensions[get_column_letter(col)].width = w

            fill = _cell_fill(cat)
            for row_i, e in enumerate(entries, start=3):
                vals = [
                    e.category, e.mechanism, e.operation,
                    e.key_bits if e.key_bits else "—",
                    _fmt_bytes(e.payload_bytes),
                    round(e.ops_per_sec, 1),
                    round(e.latency_ms, 4),
                    round(e.throughput_mbs, 2) if e.throughput_mbs else "—",
                    e.notes,
                ]
                for col, val in enumerate(vals, start=1):
                    c = ws.cell(row=row_i, column=col, value=val)
                    c.fill = fill
                    c.font = _DATA_FONT
                    c.border = _BORDER
                    c.alignment = Alignment(vertical="center")
                    if col in (6, 7, 8):
                        c.alignment = Alignment(horizontal="right", vertical="center")
                ws.row_dimensions[row_i].height = 16

            ws.freeze_panes = "A3"
            last_row = 2 + len(entries)
            ws.auto_filter.ref = f"A2:I{last_row}"

            # Embed a bar chart (ops/s per mechanism+operation)
            self._add_bar_chart(ws, entries, start_row=last_row + 3)

    def _add_bar_chart(self, ws, entries: list, start_row: int):
        """Write a small label+value table then embed a BarChart above it."""
        if not entries:
            return

        # Write label column and ops/s column for chart data
        label_col = 1
        val_col   = 2
        data_start = start_row

        ws.cell(row=data_start, column=label_col, value="Opération").font = _LABEL_FONT
        ws.cell(row=data_start, column=val_col,   value="Ops/s").font    = _LABEL_FONT

        for i, e in enumerate(entries, start=1):
            label = f"{e.mechanism} {e.operation}"
            if e.payload_bytes:
                label += f" {_fmt_bytes(e.payload_bytes)}"
            ws.cell(row=data_start + i, column=label_col, value=label)
            ws.cell(row=data_start + i, column=val_col,   value=round(e.ops_per_sec, 1))

        chart = BarChart()
        chart.type = "bar"
        chart.title = "Ops/s par opération"
        chart.y_axis.title = "Opération"
        chart.x_axis.title = "Ops/s"
        chart.legend = None
        chart.width  = 22
        chart.height = max(8, len(entries) * 0.55)

        data_ref  = Reference(ws, min_col=val_col,   min_row=data_start,
                              max_row=data_start + len(entries))
        cats_ref  = Reference(ws, min_col=label_col, min_row=data_start + 1,
                              max_row=data_start + len(entries))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        anchor_col = get_column_letter(val_col + 2)
        ws.add_chart(chart, f"{anchor_col}{data_start}")

    # -- Charts overview sheet -------------------------------------------------

    def _build_charts_sheet(self, wb: openpyxl.Workbook):
        """One sheet with a compact ops/s bar chart grouping all categories."""
        ws = wb.create_sheet(title="Vue d'ensemble")

        ws.merge_cells("A1:F1")
        tc = ws["A1"]
        tc.value = "Vue d'ensemble — Ops/s par mécanisme"
        tc.font = _TITLE_FONT
        tc.fill = _TITLE_FILL
        tc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 26

        # Compute best (max ops/s) per mechanism+operation
        best: dict[str, float] = {}
        for e in self.entries:
            key = f"{e.mechanism}\n{e.operation}"
            if e.ops_per_sec > best.get(key, 0):
                best[key] = e.ops_per_sec

        ws.cell(row=2, column=1, value="Mécanisme / Opération").font = _LABEL_FONT
        ws.cell(row=2, column=2, value="Ops/s (meilleur)").font      = _LABEL_FONT
        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 18

        for row_i, (label, ops) in enumerate(sorted(best.items()), start=3):
            ws.cell(row=row_i, column=1, value=label.replace("\n", " · "))
            ws.cell(row=row_i, column=2, value=round(ops, 1))

        last_row = 2 + len(best)
        chart = BarChart()
        chart.type    = "bar"
        chart.title   = "Performance globale — Ops/s"
        chart.y_axis.title = "Mécanisme"
        chart.x_axis.title = "Ops/s"
        chart.legend  = None
        chart.width   = 26
        chart.height  = max(12, len(best) * 0.5)

        data_ref = Reference(ws, min_col=2, min_row=2, max_row=last_row)
        cats_ref = Reference(ws, min_col=1, min_row=3, max_row=last_row)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws.add_chart(chart, "D2")
